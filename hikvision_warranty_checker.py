import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import threading
import os
import sys


class HikvisionWarrantyChecker:
    def resource_path(self, relative_path):
        """Get absolute path to resource, works for dev and for PyInstaller"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)
    
    def __init__(self, root):
        self.root = root
        self.root.title("Hikvision Warranty Checker")
        self.root.geometry("700x400")
        
        # Set icon
        try:
            icon_path = self.resource_path("icon.ico")
            self.root.iconbitmap(icon_path)
        except:
            pass
        
        self.serial_list = []
        
        # Row 1: Buttons
        button_frame = tk.Frame(root, pady=5, padx=5)
        button_frame.pack(fill=tk.X)
        
        # Load button
        self.load_btn = tk.Button(button_frame, text="Load File", command=self.load_file, width=12)
        self.load_btn.pack(side=tk.LEFT, padx=5)
        
        # Check button
        self.check_btn = tk.Button(button_frame, text="Check Warranty", command=self.check_warranty, width=15)
        self.check_btn.pack(side=tk.LEFT, padx=5)
        
        # Export button (right aligned)
        self.export_btn = tk.Button(button_frame, text="Export Excel", command=self.export_excel, width=12)
        self.export_btn.pack(side=tk.RIGHT, padx=5)
        
        # Progress label
        self.progress_label = tk.Label(button_frame, text="", font=("Arial", 8))
        self.progress_label.pack(side=tk.LEFT, padx=10)
        
        # Row 2: Listbox with Treeview
        tree_frame = tk.Frame(root, padx=10, pady=5)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview
        self.tree = ttk.Treeview(tree_frame, columns=("STT", "Serial", "Model", "Status"), 
                                  show="headings", yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        
        # Column headings
        self.tree.heading("STT", text="STT")
        self.tree.heading("Serial", text="Serial")
        self.tree.heading("Model", text="Model")
        self.tree.heading("Status", text="Trạng thái")
        
        # Column widths
        self.tree.column("STT", width=40, anchor=tk.CENTER)
        self.tree.column("Serial", width=120, anchor=tk.CENTER)
        self.tree.column("Model", width=250, anchor=tk.W)
        self.tree.column("Status", width=150, anchor=tk.CENTER)
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Footer label
        footer_label = tk.Label(root, text="Dữ liệu bảo hành được truy xuất từ lehoangcctv", 
                                font=("Arial", 8), fg="gray", anchor="w")
        footer_label.pack(anchor="w", padx=10, pady=2)
        
    def load_file(self):
        file_path = filedialog.askopenfilename(
            title="Chọn file Serial",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                
                # Clear previous data
                for item in self.tree.get_children():
                    self.tree.delete(item)
                
                self.serial_list = []
                for line in lines:
                    line = line.strip()
                    if line:
                        # Get last 9 characters
                        serial = line[-9:] if len(line) >= 9 else line
                        self.serial_list.append(serial)
                
                messagebox.showinfo("Thành công", f"Đã load {len(self.serial_list)} serial numbers")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể đọc file: {str(e)}")
    
    def check_warranty(self):
        if not self.serial_list:
            messagebox.showwarning("Cảnh báo", "Vui lòng load file serial trước!")
            return
        
        # Clear previous results
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Disable buttons during checking
        self.check_btn.config(state=tk.DISABLED)
        self.load_btn.config(state=tk.DISABLED)
        self.export_btn.config(state=tk.DISABLED)
        
        # Run in thread to avoid freezing UI
        thread = threading.Thread(target=self.check_api_thread)
        thread.daemon = True
        thread.start()
    
    def check_api_thread(self):
        total = len(self.serial_list)
        
        for idx, serial in enumerate(self.serial_list, 1):
            try:
                # Update progress
                self.root.after(0, lambda i=idx, t=total: 
                               self.progress_label.config(text=f"Đang kiểm tra: {i}/{t}"))
                
                url = f"http://sn.lehoangcctv.com:100/Checking/Checking?find={serial}"
                response = requests.get(url, timeout=10)
                
                if response.status_code == 200:
                    data = response.json()
                    if data and len(data) > 0:
                        item = data[0]
                        sn = item.get("SN", serial)
                        model = item.get("Model", "N/A")
                        status_en = item.get("Stat", "Unknown")
                        
                        # Translate status to Vietnamese
                        if "In-warranty" in status_en or "In warranty" in status_en:
                            status = "Còn bảo hành"
                        elif "Not found/Out of warranty" in status_en:
                            status = "Không có thông tin"
                        elif "Out of warranty" in status_en or "Out-of-warranty" in status_en:
                            status = "Hết bảo hành"
                        elif "Not found" in status_en:
                            status = "Không tìm thấy"
                        else:
                            status = status_en
                    else:
                        sn = serial
                        model = "Not found"
                        status = "Không tìm thấy"
                else:
                    sn = serial
                    model = "API Error"
                    status = f"Error {response.status_code}"
                    
            except Exception as e:
                sn = serial
                model = "Connection Error"
                status = str(e)
            
            # Insert to treeview
            self.root.after(0, lambda i=idx, s=sn, m=model, st=status: 
                           self.tree.insert("", tk.END, values=(i, s, m, st)))
        
        # Re-enable buttons
        self.root.after(0, self.enable_buttons)
        self.root.after(0, lambda: self.progress_label.config(text="Hoàn thành!"))
    
    def enable_buttons(self):
        self.check_btn.config(state=tk.NORMAL)
        self.load_btn.config(state=tk.NORMAL)
        self.export_btn.config(state=tk.NORMAL)
    
    def export_excel(self):
        if not self.tree.get_children():
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu để xuất!")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Lưu file Excel"
        )
        
        if file_path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Warranty Check"
                
                # Headers
                headers = ["STT", "Serial", "Model", "Trạng thái"]
                ws.append(headers)
                
                # Format headers
                header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Data
                for item in self.tree.get_children():
                    values = self.tree.item(item)["values"]
                    ws.append(values)
                
                # Adjust column widths
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 45
                ws.column_dimensions['D'].width = 20
                
                # Center align STT and Serial columns
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
                    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
                
                wb.save(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất file Excel: {file_path}")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể xuất Excel: {str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = HikvisionWarrantyChecker(root)
    root.mainloop()
